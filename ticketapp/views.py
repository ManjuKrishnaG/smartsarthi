from django.shortcuts import render,redirect,get_object_or_404
from . models import PickupRequest, CustomerDetails, CustomerAddress, ManifestResponse,ShipmentLog, ArchivedPickupRequest
from django.contrib import messages
import json
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import cache_control
from rest_framework.decorators import api_view
from rest_framework.response import Response
import requests
from logistics.views import send_manifest_details,get_bulk_reverse_manifest_status
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
import openpyxl
from django.db.models import Q
from datetime import datetime
import base64
from django.core.files.base import ContentFile

def handle_not_found(request, not_found):
    return render(request, 'ticketapp/error_400.html', status=404)

def save_api_data(api_data):
    for item in api_data:
        ticket_data = item
        customer_details_data = item.pop('customer_details', {})
        customer_address_data = customer_details_data.pop('customer_address', {})
        existing_ticket = PickupRequest.objects.filter(ticket_id=ticket_data['ticket_id']).first()
        
        # Check if the ticket with the same ID already exists
        if existing_ticket:
            # Archive the existing ticket (if required)
            archive_existing_ticket(existing_ticket)
            existing_ticket.product_name = ticket_data['product_name']
            existing_ticket.variant_no = ticket_data['variant_no']
            existing_ticket.problem_category = ticket_data['problem_category']
            existing_ticket.problem_description = ticket_data['problem_description']
            existing_ticket.date_of_purchase = ticket_data['date_of_purchase']
            existing_ticket.remarks = ticket_data.get('remarks', None)
            existing_ticket.status = 'Open'
            existing_ticket.save()
            
            # Update the associated CustomerDetails
            customer_details = existing_ticket.customerdetails
            customer_details.customer_name = customer_details_data['customer_name']
            customer_details.customer_mobile_no = customer_details_data['customer_mobile_no']
            customer_details.customer_email_id = customer_details_data['customer_email_id']
            customer_details.save()
            
            # Update the associated CustomerAddress
            customer_address = customer_details.customeraddress
            customer_address.house_no = customer_address_data['house_no']
            customer_address.address_line_1 = customer_address_data['address_line_1']
            customer_address.address_line_2 = customer_address_data.get('address_line_2', None)
            customer_address.landmark = customer_address_data.get('landmark', None)
            customer_address.city = customer_address_data['city']
            customer_address.state = customer_address_data['state']
            customer_address.pincode = customer_address_data['pincode']
            customer_address.save()
        else:
        
            # Create PickupRequest object for the new ticket
            ticket = PickupRequest.objects.create(**ticket_data)

            # Create CustomerDetails object
            customer_details = CustomerDetails.objects.create(ticket_id=ticket, **customer_details_data)

            # Create CustomerAddress object
            CustomerAddress.objects.create(customer_details=customer_details, **customer_address_data)

def archive_existing_ticket(existing_ticket):
    # Create a new archived ticket entry or update existing archived record
    # Here you can decide how to handle the archival process, whether to create a new record, update existing, or simply keep track in a separate history table.
    # For simplicity, let's assume we have an ArchivedPickupRequest model with similar fields as PickupRequest
    customer_details_data = existing_ticket.customerdetails
    customer_address_data = customer_details_data.customeraddress
    archived_ticket_data = {
        'ticket_id': existing_ticket.ticket_id,
        'product_name': existing_ticket.product_name,
        'variant_no': existing_ticket.variant_no,
        'problem_category': existing_ticket.problem_category,
        'problem_description': existing_ticket.problem_description,
        'date_of_purchase': existing_ticket.date_of_purchase,
        'invoice_copy': existing_ticket.invoice_copy,
        'video_of_issue': existing_ticket.video_of_issue,
        'front_image': existing_ticket.front_image,
        'back_image': existing_ticket.back_image,
        'customer_name': customer_details_data.customer_name,
        'customer_mobile_no': customer_details_data.customer_mobile_no,
        'customer_email_id': customer_details_data.customer_email_id,
        'house_no': customer_address_data.house_no,
        'address_line_1': customer_address_data.address_line_1,
        'address_line_2': customer_address_data.address_line_2,
        'landmark': customer_address_data.landmark,
        'city': customer_address_data.city,
        'state': customer_address_data.state,
        'pincode': customer_address_data.pincode,
        'remarks': existing_ticket.remarks,
        'created_at': existing_ticket.created_at        
    }
    ArchivedPickupRequest.objects.create(**archived_ticket_data)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@csrf_exempt
def display_ticket_list(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Define the expected structure
            expected_structure = {
                "ticket_id": str,
                "product_name": str,
                "variant_no": str,
                "problem_category": str,
                "problem_description": str,
                "date_of_purchase": str,
                "customer_details": {
                    "customer_name": str,
                    "customer_mobile_no": str,
                    "customer_email_id": str,
                    "customer_address": {
                        "house_no": str,
                        "address_line_1": str,
                        "address_line_2": str,
                        "landmark": str,
                        "city": str,
                        "state": str,
                        "pincode": str
                    }
                }
            }

            # Check if the received data matches the expected structure
            def validate_structure(data, expected_structure, path=""):
                if isinstance(expected_structure, dict):
                    if not isinstance(data, dict):
                        return False, f"{path}: Invalid data structure"
                    for key, value in expected_structure.items():
                        if key not in data:
                            return False, f"{path}{key}: is mandatory"
                        if isinstance(value, dict):
                            valid, message = validate_structure(data[key], value, f"{path}{key}.")
                            if not valid:
                                return False, message
                        elif isinstance(value, type):
                            if value == str and (not data[key] or data[key].strip() == ""):
                                return False, f"{path}{key}: is mandatory"
                            elif value != str and data[key] is None:
                                return False, f"{path}{key}: is mandatory"
                    return True, None
                elif isinstance(expected_structure, type):
                    return isinstance(data, expected_structure), None
                else:
                    return False, f"{path}: Invalid data structure"
 
            # Perform validation
            valid, message = validate_structure(data, expected_structure)
            if not valid:
                return JsonResponse({'status': 'error',  'status code': 103, 'message': message })
 
            save_api_data([data])
            return JsonResponse({'status': 'success', 'status code': 100})
        except Exception as e:
            return JsonResponse({'status': 'error', 'status code': 104, 'message': str(e)})
    else:
        return JsonResponse({'status': 'error', 'status code': 104, 'message': 'Only POST requests are allowed'})
    
#API for receiving image
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@csrf_exempt
def receive_image(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Define the expected structure
            expected_structure = {
                "CaseNumber": str,
                "KindOfFile": str,
				"FileString": str,
            }

            # Check if the received data matches the expected structure
            def validate_structure(data, expected_structure, path=""):
                if isinstance(expected_structure, dict):
                    if not isinstance(data, dict):
                        return False, f"{path}: Invalid data structure"
                    for key, value in expected_structure.items():
                        if key not in data:
                            return False, f"{path}{key}: is mandatory"
                        if isinstance(value, dict):
                            valid, message = validate_structure(data[key], value, f"{path}{key}.")
                            if not valid:
                                return False, message
                        elif isinstance(value, type):
                            if value == str and (not data[key] or data[key].strip() == ""):
                                return False, f"{path}{key}: is mandatory"
                            elif value != str and data[key] is None:
                                return False, f"{path}{key}: is mandatory"
                    return True, None
                elif isinstance(expected_structure, type):
                    return isinstance(data, expected_structure), None
                else:
                    return False, f"{path}: Invalid data structure"
 
            # Perform validation
            valid, message = validate_structure(data, expected_structure)
            if not valid:
                return JsonResponse({'status': 'error',  'status code': 103, 'message': message })
			
            save_image([data])
            return JsonResponse({'status': 'success', 'status code': 100})
        except Exception as e:
            return JsonResponse({'status': 'error', 'status code': 104, 'message': str(e)})
    else:
        return JsonResponse({'status': 'error', 'status code': 104, 'message': 'Only POST requests are allowed'})
    
def save_image(api_data):
    for item in api_data:
        case_number = item["CaseNumber"]
        kind_of_file = item["KindOfFile"]
        image_data = item["FileString"]  # Assuming this is the base64 encoded image string
              
        base64_data = image_data[len("data:image/png;base64,"):]

        # Decode the base64 string
        decoded_image_data = base64.b64decode(base64_data)
        # Check if the case number matches any of the ticket IDs
        pickup_requests = PickupRequest.objects.filter(ticket_id=case_number)
        
        if pickup_requests.exists():
            for pickup_request in pickup_requests:
                # Determine the field to save based on the kind of file
                field_to_save = None
                if kind_of_file == "Frontside Image of the watch":
                    field_to_save = 'front_image'
                    file_extension = 'png'
                elif kind_of_file == "Backside Image of the watch":
                    field_to_save = 'back_image'
                    file_extension = 'png'
                elif kind_of_file == "Video Of the Issue":
                    field_to_save = 'video_of_issue'
                    file_extension = 'mp4'
                elif kind_of_file == "Copy of the invoice":
                    field_to_save = 'invoice_copy'
                    file_extension = 'png'
                
                # Save the decoded image data to the appropriate field
                if field_to_save:
                    # Create a ContentFile from the decoded image data
                    content_file = ContentFile(decoded_image_data)

                    # Set the file name for the ContentFile
                    content_file.name = f"{case_number}_{kind_of_file}.{file_extension}"

                    # Save the ContentFile to the appropriate field
                    setattr(pickup_request, field_to_save, content_file)

                    # Ensure the model is saved after setting the file field
                    pickup_request.save()

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def index(request):
    user_groups = request.user.groups.all()
    u=request.user
    open_count = PickupRequest.objects.filter(status='Open').count()
    accepted_count = PickupRequest.objects.filter(status='Accepted').count()
    declined_count = PickupRequest.objects.filter(status='Declined').count()

    # Calculate percentages based on the total count of tickets (assuming there is a total_count variable)
    total_count = PickupRequest.objects.count()

    open_percentage = (open_count / total_count) * 100 if total_count > 0 else 0
    declined_percentage = (declined_count / total_count) * 100 if total_count > 0 else 0
    accepted_percentage = (accepted_count / total_count) * 100 if total_count > 0 else 0

    context={
        'accepted_count': accepted_count,
        'declined_count': declined_count,
        'open_count': open_count,
        'user_groups': user_groups,
        'u':u,
        'declined_percentage': declined_percentage,
        'open_percentage': open_percentage,
        'accepted_percentage': accepted_percentage,
    }
    return render(request,'ticketapp/index.html',context)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def dashboard_ticket_details(request, tkt_id):
    user_groups = request.user.groups.all()
    ticket = get_object_or_404(PickupRequest, ticket_id=tkt_id)
    context = {
        'ticket': ticket,
        'user_groups': user_groups,
    }
    return render(request, 'ticketapp/dashboard_ticket_details.html', context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def dashboard_customer_details(request, tkt_id):
    user_groups = request.user.groups.all()
    ticket = get_object_or_404(PickupRequest, ticket_id=tkt_id)
    customer = get_object_or_404(CustomerDetails, ticket_id=ticket)
    address = get_object_or_404(CustomerAddress, customer_details=customer)
    try:
        awbresponse = ManifestResponse.objects.get(ticket_id=tkt_id)
    except ManifestResponse.DoesNotExist:
        awbresponse = None
    context = {
        'customer': customer,
        'address': address,
        'ticket': ticket,
        'user_groups': user_groups,
        'awbresponse': awbresponse,
    }
    return render(request,'ticketapp/dashboard_customer_details.html',context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def ticket_reports(request):
    user_groups = request.user.groups.all()
    error_messages = []  # Initialize an empty list to store error messages
    
    if request.method == "POST":
        fromdate = request.POST.get('fromdate')
        todate = request.POST.get('todate')

        # Check if both fromdate and todate are provided
        if fromdate and todate:
            try:
                # Convert the dates to the correct format (YYYY-MM-DD)
                fromdate_formatted = datetime.strptime(fromdate, '%Y-%m-%d').date()
                todate_formatted = datetime.strptime(todate, '%Y-%m-%d').date()

                # Calculate the difference in days
                difference_in_days = (todate_formatted - fromdate_formatted).days
                
                # Check if the difference exceeds 30 days
                if difference_in_days > 30:
                    error_messages.append('The duration between fromdate and todate should not exceed 30 days.')
                else:
                    # Fetch tickets created between fromdate and todate
                    searchresult = PickupRequest.objects.filter(created_at__range=[fromdate_formatted, todate_formatted])

                    # Render the template with the filtered data
                    return render(request, 'ticketapp/data_table.html', {'tickets_list': searchresult, 'user_groups': user_groups})

            except ValueError:
                error_messages.append('Invalid date format. Please provide dates in YYYY-MM-DD format.')

        else:
            error_messages.append('Please provide both fromdate and todate.')

    # If there are any error messages, they will be passed to the template through context
    tickets_list = PickupRequest.objects.all()
    context = {
        'tickets_list': tickets_list,
        'user_groups': user_groups,
        'error_messages': error_messages,  # Pass error messages to template
    }
    return render(request, 'ticketapp/data_table.html', context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def ticket_list(request):
    user_groups = request.user.groups.all()
    ticket_list=PickupRequest.objects.all()
    context={
        'ticket_list':ticket_list,
        'user_groups': user_groups,
    }
    return render(request,'ticketapp/ticket_list.html',context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def all_ticket_list(request):
    user_groups = request.user.groups.all()
    ticket_list=PickupRequest.objects.all()
    context={
        'ticket_list':ticket_list,
        'user_groups': user_groups,
    }
    return render(request,'ticketapp/all_ticket_list.html',context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def accepted_ticket_list(request):
    user_groups = request.user.groups.all()
    ticket_list=PickupRequest.objects.all()
    context={
        'ticket_list':ticket_list,
        'user_groups': user_groups,
    }
    return render(request,'ticketapp/accepted_ticket_list.html',context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def declined_ticket_list(request):
    user_groups = request.user.groups.all()
    ticket_list=PickupRequest.objects.all()
    context={
        'ticket_list':ticket_list,
        'user_groups': user_groups,
    }
    return render(request,'ticketapp/declined_ticket_list.html',context)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def accept_ticket(request, ticket_id):
    ticket = get_object_or_404(PickupRequest, id=ticket_id)
    id= ticket.ticket_id
    send_manifest_details(request, ticket_id)
    manifest_response = ManifestResponse.objects.get(ticket_id=id)
    if manifest_response.return_code == 100:
        ticket.status = 'Accepted'
        ticket.save()
        send_status(request,ticket_id)
        if ticket.locked_by == request.user:
            ticket.locked = False
            ticket.locked_by = None
            ticket.locked_at = None
            ticket.save()
        messages.success(request, f'Ticket {ticket.ticket_id} accepted successfully.') 
    else:
        if ticket.locked_by == request.user:
            ticket.locked = False
            ticket.locked_by = None
            ticket.locked_at = None
            ticket.save()
        messages.error(request, f'Error: {manifest_response.return_message}') 
    return redirect('ticket_list')

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def decline_ticket(request, ticket_id):
    if request.method == 'POST':
        data = json.loads(request.body)
        remarks = data.get('remarks')
        print(remarks)
        ticket = get_object_or_404(PickupRequest, ticket_id=ticket_id)
        ticket.status = 'Declined'
        ticket.remarks = remarks
        ticket.save()
        messages.error(request, f'Ticket {ticket.ticket_id} declined successfully.')
        send_status(request,ticket.id)
        if ticket.locked_by == request.user:
            ticket.locked = False
            ticket.locked_by = None
            ticket.locked_at = None
            ticket.save()
        return JsonResponse({'success': True, 'message': 'Ticket declined successfully'})
    return JsonResponse({'success': False, 'error': 'Invalid request'})

from django.utils import timezone
from datetime import timedelta
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def ticket_details(request, tkt_id):
    user_groups = request.user.groups.all()
    ticket = get_object_or_404(PickupRequest, ticket_id=tkt_id)
    if ticket.locked and ticket.locked_by != request.user:
        lock_duration = timezone.now() - ticket.locked_at
        if lock_duration > timedelta(minutes=1):
            # Release the lock
            ticket.locked = True
            ticket.locked_by = request.user
            ticket.locked_at = timezone.now() 
            ticket.save()
            context = {
                'ticket': ticket,
                'user_groups': user_groups,
            }
            return render(request, 'ticketapp/Product_Details.html', context)
        
        else:
            # Redirect to another page or show a message indicating that the ticket is locked
            messages.error(request, "This ticket is currently locked by another user.")
            return redirect('ticket_list')
    
    ticket.locked = True
    ticket.locked_by = request.user
    ticket.locked_at = timezone.now() 
    ticket.save()
    context = {
        'ticket': ticket,
        'user_groups': user_groups,
    }
    return render(request, 'ticketapp/Product_Details.html', context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def customer_details(request, tkt_id):
    user_groups = request.user.groups.all()
    ticket = get_object_or_404(PickupRequest, ticket_id=tkt_id)
    customer = get_object_or_404(CustomerDetails, ticket_id=ticket)
    address = get_object_or_404(CustomerAddress, customer_details=customer)
    if ticket.locked and ticket.locked_by != request.user:
        # Redirect to another page or show a message indicating that the ticket is locked
        messages.error(request, "This ticket is currently locked by another user.")
        return redirect('ticket_list')
    
    ticket.locked = True
    ticket.locked_by = request.user
    ticket.save()
    try:
        manifest = ManifestResponse.objects.get(ticket_id=tkt_id)
    except ManifestResponse.DoesNotExist:
        manifest = None
    context = {
        'customer': customer,
        'address': address,
        'ticket': ticket,
        'user_groups': user_groups,
        'manifest': manifest,
    }
    return render(request,'ticketapp/customer_details.html',context)    

def errorpage(request):
    return render(request,'ticketapp/error_400.html')

@api_view(['GET', 'POST'])
def send_status(request,ticket_id):
    url = "https://titan-ecomm-exp-app-uat.us-e2.cloudhub.io/api/exp/case/v1/UpdateCaseStatus"
    ticket = get_object_or_404(PickupRequest, id=ticket_id)
    customer_details = get_object_or_404(CustomerDetails,ticket_id=ticket_id)
    customer_address = get_object_or_404(CustomerAddress,customer_details=customer_details)
    # Define the payload data to send to the API
    payload = {
        "CaseNumber": ticket.ticket_id,
        "CaseStatus": ticket.status,
        "CaseComments": ticket.remarks,
    }

    # Define the basic authentication credentials
    username = "Titan_Mule"
    password = "admin_t!tan_mule"
    token= "VGl0YW5fTXVsZTphZG1pbl90IXRhbl9tdWxl"
    # Define the headers with basic authentication
    headers = {
        "partnerId": "QMetrics",
        "Authorization": f"Basic {token}",
        "Content-Type": "application/json",
    }

        # Send the POST request to the API endpoint
    response = requests.post(url, headers=headers, data=json.dumps(payload))
    response_data = response.json()
    print(response_data)
    return Response(response_data)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def track_package(request,awb_no):
    user_groups = request.user.groups.all()
    if ManifestResponse.objects.filter(awb_number=awb_no).exists():
        try:
            awbresponse = get_object_or_404(ManifestResponse, awb_number=awb_no)
            response = get_bulk_reverse_manifest_status(request,awb_no)
            shipment_response = ShipmentLog.objects.get(awb_number=awb_no)
            context = {
                'user_groups': user_groups,
                'awbresponse': awbresponse,
                'shipment_response': shipment_response,
            }
            return render(request, 'ticketapp/track_package.html', context)
        except ShipmentLog.DoesNotExist:
            # Handle the case where ShipmentLog does not exist for the given AWB number
            return render(request, 'ticketapp/track_package.html', {'error_message': 'Shipment data not found.'})
    else:
        # Render a different view if AWB number does not exist
        return render(request, 'ticketapp/track.html', {'awb_no': awb_no, 'user_groups': user_groups, 'error_message': 'AWB number not found.'})
    


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def track(request):
    user_groups = request.user.groups.all()
    context = {
        'user_groups': user_groups
    }
    return render(request,'ticketapp/track.html',context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def profile(request):
    user_groups = request.user.groups.all() 
    user = request.user
    user_name = user.username  
    user_role = user.groups.first().name
    context = {
        'user_name': user_name,
        'user_role': user_role,
        'user_groups': user_groups,
    }
    return render(request, 'ticketapp/profile.html', context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def download_excel(request):
    tickets_list = PickupRequest.objects.all()
    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
 
    # Define the header row with bold and colored text
    header_row = ['Ticket id', 'Product name', 'Variant no', 'Problem category', 'Date of purchase', 'Status']
    for col_num, col_name in enumerate(header_row, start=1):
        ws.cell(row=1, column=col_num, value=col_name)
        ws.cell(row=1, column=col_num).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=1, column=col_num).fill = openpyxl.styles.PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  # Pink color
 
    # Write data rows
    for row_num, ticket in enumerate(tickets_list, start=2):
        ws.append([ticket.ticket_id, ticket.product_name, ticket.variant_no, ticket.problem_category, ticket.date_of_purchase, ticket.status])
 
    # Create the HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="tickets.xlsx"'
 
    # Save the workbook to the response
    wb.save(response)
 
    return response

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def download_pdf(request):
    tickets_list = PickupRequest.objects.all()
 
    template_path = 'ticketapp/pdf_template.html'
    context = {'tickets_list': tickets_list}
    # Create a Django response object
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="tickets.pdf"'
 
    # Render the template to a HTML string
    template = get_template(template_path)
    html = template.render(context)
 
    # Create PDF
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def search_ticket(request):
    user_groups = request.user.groups.all()
    ticket_id = request.GET.get('ticket_id')  # Get the ticket_id from the request
   
    # Filter PickupRequest objects based on the provided ticket_id
    if ticket_id:
        ticket_list = PickupRequest.objects.filter(Q(ticket_id__icontains=ticket_id) | Q(new_ticket_id__icontains=ticket_id))
    else:
        ticket_list = None
   
    context = {                
        'user_groups': user_groups,
        'ticket_list': ticket_list,
    }
           
    return render(request, 'ticketapp/search.html', context)


from django.http import HttpResponse
from django.conf import settings
import os

def serve_media_file(request, file_name):
    # Construct the absolute path to the requested file
    file_path = os.path.join(settings.MEDIA_ROOT, 'attachments', file_name)
    try:
        # Open the file in binary mode and read its contents
        with open(file_path, 'rb') as f:
            file_data = f.read()
        # Determine the appropriate MIME type based on the file extension
        content_type = 'image/jpeg' if file_name.endswith('.png') else 'video/mp4'
        # Return the file data with the appropriate content type
        return HttpResponse(file_data, content_type=content_type)
    except FileNotFoundError:
        # Return a 404 Not Found response if the file is not found
        return HttpResponse(status=404)